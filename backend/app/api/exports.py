"""Download Centre — formatted XLSX datasets + a branded PDF executive summary.

Export 1: Buyer Intelligence Dataset (.xlsx)
Export 2: Project Dataset with risk flags (.xlsx)
Export 3: Buyer-Project Mapping with sources + confidence (.xlsx)
Export 4: Executive Summary (.pdf, Xynteo-branded)
"""
from __future__ import annotations

import io

from fastapi import APIRouter, Depends, Response
from sqlalchemy.orm import Session, joinedload

from app.db.models import BuyerProjectLink, RiskFlag
from app.db.session import get_db
from app.schemas import ProjectFilters
from app.services import analytics as analytics_svc
from app.services import filters as filter_svc

router = APIRouter(prefix="/exports", tags=["exports"])

XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# --- Xynteo brand palette (from xynteo.com) ---
NAVY = "001E3C"
BLUE = "0A5AD7"
LIGHTBLUE = "2DAFE6"
GREEN = "00C873"
MAGENTA = "FF0064"


def _xlsx(sheet_title: str, columns: list[str], rows: list[list], filename: str) -> Response:
    """Build a readable .xlsx: bold navy header, frozen header row, autofilter, sensible column widths."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    ws.append(columns)
    for c, _ in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 22

    for r in rows:
        ws.append(r)

    # Column widths from content (capped so long text/URLs don't blow the layout), wrap long cells.
    for c, col_name in enumerate(columns, start=1):
        longest = len(str(col_name))
        for r in rows:
            v = r[c - 1] if c - 1 < len(r) else ""
            longest = max(longest, len(str(v if v is not None else "")))
        width = min(max(longest + 2, 12), 55)
        ws.column_dimensions[get_column_letter(c)].width = width
        if col_name.lower() in ("evidence", "source_url", "volume_basis", "risk_description", "profile_summary"):
            for r_idx in range(2, len(rows) + 2):
                ws.cell(row=r_idx, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return Response(content=buf.getvalue(), media_type=XLSX_MEDIA,
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.post("/buyers.xlsx")
def export_buyers(f: ProjectFilters, db: Session = Depends(get_db)):
    dash = analytics_svc.build_dashboard(db, f)
    seen, buyers = set(), []
    for b in sorted(dash.top_buyers + dash.repeat_buyers, key=lambda x: x.total_estimated_volume, reverse=True):
        if b.name in seen:
            continue
        seen.add(b.name)
        buyers.append(b)
    columns = ["Buyer", "Industry", "Entity type", "Confidence", "SBTi status", "SBTi alignment",
               "Est. volume (tCO2e)", "Retired (tCO2e)", "Non-retired (tCO2e)", "Projects",
               "Countries", "Project types", "Purchase years", "Repeat buyer", "Repeat score", "HQ country"]
    rows = [[b.name, b.industry, b.entity_type, b.confidence_tier, b.sbti_status, b.sbti_alignment,
             round(b.total_estimated_volume), round(b.total_retired_volume), round(b.total_non_retired_volume),
             b.num_projects, b.num_countries, b.num_project_types, b.num_purchase_years,
             "Yes" if b.is_repeat_buyer else "No", b.repeat_buyer_score, b.hq_country or ""] for b in buyers]
    return _xlsx("Buyer Intelligence", columns, rows, "buyer_intelligence.xlsx")


@router.post("/projects.xlsx")
def export_projects(f: ProjectFilters, db: Session = Depends(get_db)):
    projects = filter_svc.query_projects(db, f).all()
    pid_set = {p.id for p in projects}
    risk_map: dict[int, list[RiskFlag]] = {}
    for rk in (db.query(RiskFlag).filter(RiskFlag.project_id.in_(pid_set)).all() if pid_set else []):
        risk_map.setdefault(rk.project_id, []).append(rk)

    def primary(pid):
        rs = risk_map.get(pid, [])
        if not rs:
            return "", ""
        top = max(rs, key=lambda r: r.severity_score)
        return top.risk_category.replace("_", " ").title(), int(top.severity_score)

    columns = ["Project ID", "Project name", "Registry", "Type", "Reduction/Removal", "Country",
               "Region", "Status", "Vintage", "Credits issued", "Credits retired", "Credits remaining",
               "Developer", "Eligible", "Primary risk", "Risk severity", "All risk flags"]
    rows = []
    for p in projects:
        pr_cat, pr_sev = primary(p.id)
        rows.append([p.project_id, p.project_name, p.registry, p.type, p.reduction_removal, p.country,
                     p.region, p.voluntary_status, p.first_vintage_year or "", round(p.credits_issued),
                     round(p.credits_retired), round(p.credits_remaining), p.developer,
                     "Yes" if p.is_eligible else "No", pr_cat, pr_sev,
                     "; ".join(f"{r.risk_category}({int(r.severity_score)})" for r in risk_map.get(p.id, []))])
    return _xlsx("Projects", columns, rows, "project_dataset.xlsx")


@router.post("/buyer-project-mapping.xlsx")
def export_mapping(f: ProjectFilters, db: Session = Depends(get_db)):
    projects = filter_svc.query_projects(db, f).all()
    pid_set = {p.id for p in projects}
    links = (db.query(BuyerProjectLink)
             .options(joinedload(BuyerProjectLink.buyer), joinedload(BuyerProjectLink.project))
             .filter(BuyerProjectLink.project_id.in_(pid_set)).all()) if pid_set else []
    columns = ["Project ID", "Project name", "Buyer", "Buyer role", "Transaction type",
               "Est. volume (tCO2e)", "Purchase year", "Confidence", "Confidence score", "Verdict",
               "Source type", "Source URL", "Evidence"]
    rows = [[
        l.project.project_id if l.project else "", l.project.project_name if l.project else "",
        l.buyer.name if l.buyer else "", l.buyer_role, l.transaction_type,
        round(l.estimated_volume_tco2e) if l.estimated_volume_tco2e else "", l.purchase_year or "",
        l.confidence_tier, round(l.confidence_score), l.verdict or "", l.source_type,
        l.source_url, l.evidence_summary,
    ] for l in sorted(links, key=lambda x: x.confidence_score, reverse=True)]
    return _xlsx("Buyer-Project Mapping", columns, rows, "buyer_project_mapping.xlsx")


def _ascii(s: str) -> str:
    """fpdf core fonts are latin-1 only — replace common unicode with safe equivalents."""
    if s is None:
        return ""
    repl = {"₂": "2", "·": "-", "—": "-", "–": "-", "≥": ">=", "→": "->", "’": "'", "“": '"', "”": '"', "•": "-"}
    for k, v in repl.items():
        s = s.replace(k, v)
    return s.encode("latin-1", "replace").decode("latin-1")


@router.post("/executive-summary.pdf")
def export_exec_summary(f: ProjectFilters, db: Session = Depends(get_db)):
    from fpdf import FPDF

    dash = analytics_svc.build_dashboard(db, f)
    k = dash.kpis
    seg = f"{f.country or 'All countries'} - {f.project_type or 'All project types'}"

    def rgb(hexs):
        return tuple(int(hexs[i:i + 2], 16) for i in (0, 2, 4))

    navy, blue, lightblue, magenta, ink = rgb(NAVY), rgb(BLUE), rgb(LIGHTBLUE), rgb(MAGENTA), (26, 26, 26)

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(True, margin=15)
    pdf.add_page()

    # Header band
    pdf.set_fill_color(*navy)
    pdf.rect(0, 0, 210, 30, "F")
    pdf.set_text_color(255, 255, 255)
    pdf.set_xy(14, 8)
    pdf.set_font("Helvetica", "B", 19)
    pdf.cell(0, 9, "Carbon Credit Buyer Intelligence")
    pdf.set_xy(14, 18)
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 7, _ascii(f"Executive Summary  -  {seg}"))
    pdf.ln(24)

    # KPI strip
    pdf.set_text_color(*ink)
    kpis = [("Buyers", f"{k.total_buyers}"), ("Est. volume (tCO2e)", f"{k.total_estimated_volume:,.0f}"),
            ("Projects", f"{k.total_projects}"), ("Repeat %", f"{k.repeat_buyer_pct:.0f}%"),
            ("SBTi-aligned %", f"{k.sbti_aligned_pct:.0f}%")]
    w = 182 / len(kpis)
    x0 = 14
    for label, val in kpis:
        pdf.set_fill_color(240, 244, 250)
        pdf.rect(x0, pdf.get_y(), w - 3, 20, "F")
        pdf.set_xy(x0, pdf.get_y() + 3)
        pdf.set_text_color(*blue)
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(w - 3, 7, _ascii(val), align="C")
        pdf.set_xy(x0, pdf.get_y() + 8)
        pdf.set_text_color(90, 90, 90)
        pdf.set_font("Helvetica", "", 7.5)
        pdf.multi_cell(w - 3, 3.4, _ascii(label), align="C")
        x0 += w
    pdf.ln(26)

    def heading(t):
        pdf.set_text_color(*navy)
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 8, _ascii(t), new_x="LMARGIN", new_y="NEXT")
        pdf.set_draw_color(*lightblue)
        pdf.set_line_width(0.6)
        y = pdf.get_y()
        pdf.line(14, y, 196, y)
        pdf.ln(2)

    # Market overview
    heading("Market Overview")
    pdf.set_text_color(*ink)
    pdf.set_font("Helvetica", "", 10)
    pdf.multi_cell(182, 5, _ascii(
        f"This segment covers {k.total_projects} eligible offset projects. Buyer intelligence is compiled from "
        "public retirement disclosures, registry records, corporate & ESG reports, press releases and market "
        "databases - every buyer and risk carries a source and confidence tier."))
    pdf.ln(3)

    # Top buyers table
    heading("Top Buyers")
    cols = [(10, "#"), (66, "Buyer"), (40, "Industry"), (34, "Est. tCO2e"), (16, "Conf."), (16, "SBTi")]
    pdf.set_font("Helvetica", "B", 8.5)
    pdf.set_fill_color(*blue)
    pdf.set_text_color(255, 255, 255)
    for wdt, name in cols:
        pdf.cell(wdt, 7, _ascii(name), border=0, fill=True, align="L")
    pdf.ln(7)
    pdf.set_text_color(*ink)
    pdf.set_font("Helvetica", "", 8.5)
    top = dash.top_buyers[:12]
    if not top:
        pdf.cell(0, 6, "No buyers identified for this segment yet.", new_x="LMARGIN", new_y="NEXT")
    for i, b in enumerate(top, 1):
        fill = i % 2 == 0
        pdf.set_fill_color(244, 247, 251)
        vals = [str(i), b.name[:40], (b.industry or "")[:22],
                f"{b.total_estimated_volume:,.0f}" if b.total_estimated_volume else "-",
                b.confidence_tier, b.sbti_alignment.replace("SBTi ", "").replace("Not ", "No")[:6]]
        for (wdt, _), v in zip(cols, vals):
            pdf.cell(wdt, 6, _ascii(v), border=0, fill=fill)
        pdf.ln(6)
    pdf.ln(3)

    # Key risks
    heading("Key Risks")
    pdf.set_font("Helvetica", "", 9)
    risks = sorted(dash.risks, key=lambda x: x.severity_score, reverse=True)[:8]
    if not risks:
        pdf.cell(0, 6, "No material red flags surfaced.", new_x="LMARGIN", new_y="NEXT")
    for r in risks:
        sev = int(r.severity_score)
        pdf.set_text_color(*(magenta if sev >= 70 else (200, 120, 0) if sev >= 50 else (110, 110, 110)))
        pdf.set_font("Helvetica", "B", 9)
        cat = r.risk_category.replace("_", " ").title()
        pdf.multi_cell(182, 4.6, _ascii(f"[{sev}] {cat}"))
        pdf.set_text_color(*ink)
        pdf.set_font("Helvetica", "", 9)
        pdf.multi_cell(182, 4.6, _ascii(r.ai_summary))
        pdf.ln(1.5)

    pdf.set_y(-14)
    pdf.set_text_color(140, 140, 140)
    pdf.set_font("Helvetica", "I", 7.5)
    pdf.multi_cell(182, 3.5, _ascii(
        "Generated by the Carbon Credit Buyer Intelligence Platform. Every buyer claim is traceable to a cited "
        "source and confidence score in the Buyer-Project Mapping export. Project data aggregated from public "
        "voluntary carbon registries."))

    # fpdf2 returns a bytearray; starlette's Response only passes through `bytes`
    # (anything else it tries to .encode()), so coerce explicitly.
    data = bytes(pdf.output())
    return Response(content=data, media_type="application/pdf",
                    headers={"Content-Disposition": 'attachment; filename="executive_summary.pdf"'})
